# -*- coding: utf-8 -*-
import sys
sys.path.append('./entity')
import os
from  entity.GISEntity import gisSession, Area, AreaFeature,AreaRelease,Territory
from  entity.addressEntity import StandardAddrDetail, StandardAddr,addressSession
from  entity.authEntity import authSession,Org
from  entity.G7Entity import session as G7sesssion,Station

from tools import log, codeGenerator,hoauGd2bd,hoauGdBorder2bdBorder,getmaxminlonlat
from openpyxl import load_workbook
from sqlalchemy import func,and_,or_


class TransferGisData(object):
    """
       将G7数据转换为GIS数据
    """
    def __init__(self):
        # 地址
        self.__folder__ = '../data'
        #版图编码
        self.all_territory_code='244423365703815168'  #全部版图编码
        self.pick_territory_code='244423366521704448'  #自提版图编码
        self.deliver_territory_code ='244423366479761408'  #派送版图编码
             
    def filter_town(self,town_filter):
        '''
            查询4级地址编码,query_addr_detail_filter方法专用
        '''
        if  town_filter==None or town_filter.strip()=='%%' or town_filter.strip()=='':
            results = addressSession.query(StandardAddrDetail).filter(StandardAddrDetail.sname==None)
        else:
            results = addressSession.query(StandardAddrDetail).filter(
                    StandardAddrDetail.sname.like(town_filter)                   
                )
        return results

    def query_addr_detail(self, province_name,city_name,country_name,town_name=None):
        '''
            根据省、市、区县、乡镇获取地址编码
        '''
        #返回地址码，缺省为空
        code_list = []

        #去掉4级地址前后空格
        province_name = province_name.strip()
        city_name = city_name.strip()
        country_name = country_name.strip()
        town_filter=''
        if town_name!=None:
            town_filter = town_name.strip()

        #严格匹配省、市、区县、乡镇地址，一字不差                       
        results = self.query_addr_detail_filter(province_name,city_name,country_name,town_filter)
        
        #区县like匹配，乡镇严格匹配
        if results.count()==0:
            #print('区县like匹配，乡镇严格匹配')
            results = self.query_addr_detail_filter(province_name,city_name,'%'+country_name+'%',town_filter)

        #区县严格匹配，乡镇like匹配
        if results.count()==0:
            #print('区县严格匹配，乡镇like匹配')
            results = self.query_addr_detail_filter(province_name,city_name,country_name,'%'+town_filter+'%')
            
        #区县like匹配,乡镇like匹配
        if results.count()==0:
            #print('区县like匹配,乡镇like匹配')
            results = self.query_addr_detail_filter(province_name,city_name,'%'+country_name+'%','%'+town_filter+'%')

        #区县严格匹配，乡镇宽松匹配(例如'XX镇'可以匹配'XX街道')
        if results.count()==0:
            #print("区县严格匹配，乡镇宽松匹配(例如'XX镇'可以匹配'XX街道')")
            results = self.query_addr_detail_filter(province_name,city_name,country_name,town_filter[0:-1]+'%')
            
        #区县like匹配，乡镇宽松匹配(例如'XX镇'可以匹配'XX街道')
        if results.count()==0:
            #print("区县like匹配，乡镇宽松匹配(例如'XX镇'可以匹配'XX街道')")
            results = self.query_addr_detail_filter(province_name,city_name,'%'+country_name+'%',town_filter[0:-1]+'%')

        #区县宽松匹配(XX区匹配XX县)，乡镇严格匹配
        if results.count()==0:
            #print('区县宽松匹配(XX区匹配XX县)，乡镇严格匹配')
            results = self.query_addr_detail_filter(province_name,city_name,country_name[0:-1]+'%',town_filter)
            
        #区县宽松匹配(XX区匹配XX县),乡镇like匹配
        if results.count()==0:
            #print('区县宽松匹配(XX区匹配XX县),乡镇like匹配')
            results = self.query_addr_detail_filter(province_name,city_name,country_name[0:-1]+'%','%'+town_filter+'%')
            
        #记录获取的结果
        for r in results:
                code_list.append(r.standard_addr_code)

        gisSession.close()
        return code_list

    def query_addr_detail_filter(self, province_name,city_name,country_filter,town_filter=None):
        '''
            根据省、市、区县、乡镇获取地址编码的查询条件
        '''
        #处理三个直辖市
        if province_name in ['北京','北京市','天津','天津市','上海','上海市','重庆','重庆市']:
            #处理直辖市没有地级市的情况
            results = self.filter_town(town_filter).filter(
            and_(
                StandardAddrDetail.pname.like(province_name+'%'),
                StandardAddrDetail.dname.like(country_filter)           
                )
            )
        elif city_name in ['东莞','东莞市','中山','中山市','儋州','儋州市']:
            #处理没有区县的地级市的情况
            results = self.filter_town(town_filter).filter(
                and_(
                    StandardAddrDetail.pname.like(province_name+'%'),
                    StandardAddrDetail.cname.like(city_name+'%'),
                )
            )
        elif province_name in ['海南','海南省','新疆','新疆维吾尔自治区','湖北','湖北省','河南','河南省'] and city_name in ['海南','海南省','新疆','新疆维吾尔自治区','湖北','湖北省','河南','河南省','省直辖县级行政区划','']:
            #处理海南省,新疆,湖北,河南直辖的区县情况
            results = self.filter_town(town_filter).filter(
                and_(
                    StandardAddrDetail.pname.like(province_name+'%'),
                    StandardAddrDetail.dname.like(country_filter)
                )
            )
        else:     
            #根据省、市、区县、乡镇名称模糊查询数据库的地址编码
            results = self.filter_town(town_filter).filter(
            and_(
                StandardAddrDetail.pname.like(province_name+'%'),
                StandardAddrDetail.cname.like(city_name+'%'),
                StandardAddrDetail.dname.like(country_filter)
                )
            )
        return results

    def query_org(self, short_name):
        '''
            根据机构简称查询机构信息
        '''
        org_list=[]
        #去掉简称的空格
        short_name = short_name.strip()
        
        #根据简称查询机构信息
        results = authSession.query(Org).filter_by(short_name=short_name).all()

        for r in results:
            org_list.append(r)
            
        authSession.close()
        return org_list

    def query_station(self, short_name,maptype_code):
        '''
            根据机构简称查询原生产数据的制图范围数据
        '''
        station_list=[]
        tmp_short_name = short_name.strip()[1:]  #去掉简称前的N和前后空格
        
        #根据省、市、区县、乡镇查询
        results = G7sesssion.query(Station).filter_by(name=tmp_short_name).filter_by(maptype=maptype_code).all()
        for r in results:
            #只找到一条记录
            station_list.append(r)

        G7sesssion.close()
        return station_list

    def query_parent(self, province, node,territory_code):
        '''
            根据根据省份和机构类型，查询上级机构的区域编码
        '''
        #查询省级的根节点
        parent_code = None
        province_area = gisSession.query(Area).filter_by(territory_code=territory_code).filter_by(area_name=province).first()
        if province_area!=None:
            node_area = gisSession.query(Area).filter_by(territory_code=territory_code).filter_by(parent_area=province_area.area_code).filter_by(area_name=node).first()
            if node_area!=None:
                parent_code=node_area.area_code
            else:
                log('省份%s没有虚拟节点类型%s' %(province, node))
        else:
            log('没有省份根节点%s' % province)
        return parent_code

    def add(self, station, territory_code,short_name,province_name,city_name,country_name,town_name, parent_code,address):
        '''
            在数据库中增加一组制图范围数据
        '''
        #将审核状态转换为GIS数据库的枚举值
        if station.orgcheck=='1':
            status_value = 'VALID'  #已审核
        else:
            status_value = 'NEW'  #待审核

        #查询地址编码
        standard_addr_code = None
        standard_addr_code_list = self.query_addr_detail(province_name,city_name,country_name,town_name)
        if len(standard_addr_code_list)==0:
            log('机构%s没有获取到对应的行政区域编码' %short_name)
        elif len(standard_addr_code_list)==1:
            standard_addr_code=standard_addr_code_list[0]
        else:
            log('机构%s,%s,%s,%s,%s,匹配多个行政区域编码%s' %(short_name,province_name,city_name,country_name,town_name,standard_addr_code_list))           
        
        #根据机构简称获取机构的代码
        org=None
        org_list = self.query_org(short_name)
        if len(org_list)==0:
            log('无法根据简称%s查询到机构数据' % short_name)
        elif len(org_list)==1:
            org = org_list[0]
        else:
            log('根据简称%s可以查询到第%s条机构数据' % (short_name,len(org_list)))
            
        #插入区域库
        areaObj = Area(
            area_code = codeGenerator() ,  #区域编码
            area_name = short_name,  #区域名称
            parent_area = parent_code,    #上级区域
            territory_code = territory_code, #所属版图
            area_type = 'ORG',  #区域类型
            conn_entity = org.code, #关联实体编码
            standard_addr_code = standard_addr_code,   #标准地址编码
            detail_add = address,  #详细地址
            dead_sts = 'ENABLED',  #停用启用状态
            status = status_value,   #状态
            remark = station.areadesc #备注
            )
        gisSession.add(areaObj)
        gisSession.commit()

        #判断区域服务的类型
        if territory_code ==self.pick_territory_code:        
            service_type_code = 'CUSTOMER_PICK_UP'   #自提版图
        elif territory_code == self.deliver_territory_code:
            service_type_code = 'SITE_DELIVERY'  #派送版图
        elif territory_code == self.all_territory_code:  
            service_type_code = 'CARGO_COLLECT'  #全部门店
        else:
            service_type_code =''

        #百度坐标转换为高德坐标
        gd_lng,gd_lat = station.lng,station.lat
        bd_lng,bd_lat = hoauGd2bd(float(station.lng),float(station.lat)) #转换坐标点
        
        #插入区域要素和区域要素发布表
        if territory_code ==self.pick_territory_code:   
            #自提服务，只插入坐标点数据
            self.insertPoint(areaObj.area_code,service_type_code,gd_lng,gd_lat)
            if status_value== 'VALID':
                self.insertPointRelease(areaObj.area_code,service_type_code, org.code,org.name,org.short_name,org.numb,org.mobile_phone,org.telephone,gd_lng,gd_lat,bd_lng,bd_lat)

        elif territory_code == self.deliver_territory_code:
            #百度坐标范围转换为高德坐标范围
            gd_bound = station.bound
            bd_bound = hoauGdBorder2bdBorder(station.bound)  
            #派送服务，只插入坐标范围数据
            self.insertArea(areaObj.area_code, service_type_code,gd_lng,gd_lat,station.color,gd_bound)
            if status_value== 'VALID':
                #计算出高德坐标系的最大、最小经纬度范围
                min_lng,min_lat, max_lng,max_lat= getmaxminlonlat(gd_bound)        
                self.insertAreaRelease(areaObj.area_code,service_type_code, org.code,org.name,org.short_name,org.numb,org.mobile_phone,org.telephone,gd_lng,
                                       gd_lat,station.color,gd_bound,bd_lng,bd_lat,bd_bound,min_lng,min_lat,max_lng,max_lat)
               
        elif territory_code == self.all_territory_code: 
            #收件服务，只插入坐标点数据
            self.insertPoint(areaObj.area_code,service_type_code,gd_lng,gd_lat)
          
        else:
            log('%s的版图无法处理,类型为%s' % (old_recorde.name,old_recorde.maptype))

        print(short_name,province_name,city_name,country_name,town_name, parent_code)

    def insertPoint(self, area_code,service_type,lng,lat):
        '''
            插入类型为单点的服务要素
        '''
        areaFeatureObj = AreaFeature(
            code = codeGenerator(),  #区域要素编码
            area_code = area_code,  #关联区域编码
            area_lng = lng  ,#经度
            area_lat = lat  ,#经度
            service_type = service_type, #服务类型
            coordinate_system = 'GCJ02',  #坐标系
            feature_type = 'POINT',    #要素类型，代表点或者面
            )
        gisSession.add(areaFeatureObj)
        gisSession.commit()

    def insertPointRelease(self, area_code,service_type_code, org_code,org_name,org_short_name,org_numb,org_mobile_phone,org_telephone,lng,lat,bd_lng,bd_lat):
        '''
            插入类型为单点的服务要素
        '''
        #插入区域要素发布表
        areaReleaseObj=AreaRelease(
            area_code = area_code, #区域编码
            service_type = service_type_code,     #服务类型
            org_code = org_code, #机构编码
            org_name = org_name,    #机构名称
            org_short_name = org_short_name,    #机构简称
            org_numb = org_numb,   #机构代码
            org_mobile_phone = org_mobile_phone,    #手机
            org_telephone = org_telephone,    #固定电话
            coordinate_system = 'GCJ02',  #坐标系
            feature_type = 'POINT',    #要素类型，代表点或者面
            area_lng = lng,  #经度
            area_lat =  lat,  #纬度
            bd_area_lng = bd_lng,  #经度
            bd_area_lat =  bd_lat  #纬度
            )
        gisSession.add(areaReleaseObj)
        gisSession.commit()

    def insertArea(self, area_code, service_type,lng,lat,layer_color,border):
        '''
            插入类型为区域的服务要素
        '''
        areaFeatureObj = AreaFeature(
            code = codeGenerator(),  #区域服务要素编码
            area_code = area_code,  #关联区域编码
            area_lng = lng  ,#经度
            area_lat = lat  ,#经度
            service_type = service_type, #服务类型
            coordinate_system = 'GCJ02',  #坐标系
            feature_type = 'AREA',    #要素类型，代表点或者面
            layer_color = layer_color,  #图层颜色
            border = border   #坐标范围
            )
        gisSession.add(areaFeatureObj)
        gisSession.commit()

    def insertAreaRelease(self, area_code,service_type_code,org_code,org_name,org_short_name,org_numb,org_mobile_phone,
                          org_telephone,lng,lat,layer_color,border,bd_lng,bd_lat,bd_border,min_lng,min_lat,max_lng,max_lat):
        '''
            插入类型为单点的服务要素
        '''
        #插入区域要素发布表
        areaReleaseObj=AreaRelease(
            area_code = area_code ,  #区域编码
            service_type = service_type_code,     #服务类型
            org_code = org_code, #机构编码
            org_name = org_name,    #机构名称
            org_short_name = org_short_name,    #机构简称
            org_numb = org_numb,   #机构代码
            org_mobile_phone = org_mobile_phone,    #手机
            org_telephone = org_telephone,    #固定电话
            coordinate_system = 'GCJ02',  #坐标系
            feature_type = 'AREA',    #要素类型，代表点或者面
            area_lng = lng,  #经度
            area_lat =  lat,  #纬度
            layer_color = layer_color,      #图层颜色
            border = border,      #范围坐标
            bd_area_lng = bd_lng,  #经度
            bd_area_lat =  bd_lat,  #纬度
            bd_border = bd_border,      #范围坐标
            min_lng = min_lng,  #最小经度
            min_lat = min_lat,  #最小纬度
            max_lng = max_lng,  #最大经度
            max_lat = max_lat  #最大纬度
            )
        gisSession.add(areaReleaseObj)
        gisSession.commit()

    def todata(self, maptype,filename,search_maptype=None):
        '''
            将指定版图的旧数据转换为新格式数据
        '''
        #获取excel对象
        filepath= os.path.join(self.__folder__, filename)
        wb = load_workbook(filepath)
        if maptype=='派送':
            sheet = wb['派送']
            maptype_code = '2910'   #旧数据派送版图的字典值
            territory_code =self.deliver_territory_code  #新数据派送版图的编码
        elif maptype=='自提':
            sheet = wb['自提']
            maptype_code = '2911'   #旧数据自提版图的字典值
            territory_code =self.pick_territory_code   #新数据自提版图的编码
        elif maptype=='收件':    
            sheet = wb['全部版图']
            maptype_code = '2913'   #旧数据全部版图的字典值
            territory_code =self.all_territory_code   #新数据全部版图的编码

        #如果指定版图，则在指定版图里提取数据
        if search_maptype!=None:
            maptype_code= search_maptype

        #遍历所有数据
        n = 1
        for row in sheet.rows:
            if n > 1 and (row[0].value !=None):
                tree_root = row[0].value  #树结构的省份根名称
                org_type = row[1].value   #平台、门店、偏线                    
                short_name = row[2].value    #机构简称
                province_name = row[3].value   #省
                city_name = row[4].value   #市
                country_name = row[5].value  #县
                town_name = row[6].value   #镇
                address = row[7].value   #详细地址
               
                #根据省份根节点、虚拟结构、版图编码查找父级区域编码
                parent_code = self.query_parent(tree_root,org_type,territory_code)

                #根据机构简称查询旧生成数据的制图范围
                station=None
                station_list = self.query_station(short_name,maptype_code)
                if len(station_list)==0:
                    log('无法根据简称%s查询到原生产数据' % short_name)
                elif len(station_list)==1:
                    station = station_list[0]
                else:
                    log('根据简称%s可以查询到第%s条生产数据' % (short_name,len(station_list)))
                    
                if station!=None:
                    #插入记录                
                    self.add(station=station,
                             territory_code=territory_code,
                             short_name=short_name,
                             province_name=province_name,
                             city_name=city_name,
                             country_name=country_name,
                             town_name=town_name,
                             parent_code =parent_code,
                             address=address
                             )               
            n = n + 1

            
    def initTerritory(self):
        '''
            初始化版图数据
        '''
        name_dict = {'CARGO_COLLECT':'收件版图', 'SITE_DELIVERY':'派送版图', 'CUSTOMER_PICK_UP':'自提版图'}
        n=0
        for key in name_dict:
            obj = Territory(
                territory_code = codeGenerator() ,  #版图编码
                territory_name = name_dict[key],  #版图名称
                service_type = key,     #服务类型                
                show_order = n    #显示顺序
                )
            gisSession.add(obj)
            gisSession.commit()
            n +=1

    def initData(self, territory):
        '''
            初始化每个版图下的树形结构
        '''
        if territory=='派送':
            territory_code =self.deliver_territory_code  #新数据派送版图的编码
        elif territory=='自提':
            territory_code =self.pick_territory_code   #新数据自提版图的编码
        else:    
            territory_code =self.all_territory_code   #新数据全部版图的编码
        province_list = [
            '北京',
            '天津',
            '河北',
            '山西',
            '内蒙古',
            '辽宁',
            '吉林',
            '黑龙江',
            '上海',
            '江苏',
            '浙江',
            '安徽',
            '福建',
            '江西',
            '山东',
            '河南',
            '湖北',
            '湖南',
            '广东',
            '广西',
            '海南',
            '重庆',
            '四川',
            '贵州',
            '云南',
            '西藏',
            '陕西',
            '甘肃',
            '青海',
            '宁夏',
            '新疆',
            '香港',
            '澳门',
            '台湾'
        ]
        node_list =['门店','平台','偏线'] 

        for province in province_list:
            #查询省份的标准地址编码
            adcode = self.query_adcode(province)
            #插入省份数据
            areaObj = Area(
                area_code = codeGenerator() ,  #区域编码
                area_name = province,  #区域名称
                parent_area = '0',    #上级区域
                territory_code = territory_code, #所属版图
                area_type = 'STANDARD_ADDRESS',  #区域类型
                conn_entity = adcode,   #省份的关联实体填写对应的标准地址码
                dead_sts = 'ENABLED',  #停用启用状态
                status  = 'VALID'    #状态                
                 )
            gisSession.add(areaObj)
            gisSession.commit()

            #插入省份的下属虚拟机构
            for node in node_list:
                areaObj_node = Area(
                    area_code = codeGenerator() ,  #区域编码
                    area_name = node,  #区域名称
                    parent_area = areaObj.area_code,    #上级区域
                    territory_code = territory_code, #所属版图
                    area_type = 'VIRTUAL_NODE',  #区域类型
                    dead_sts = 'ENABLED',  #停用启用状态
                    status  = 'VALID'    #状态      
                     )
                gisSession.add(areaObj_node)
                gisSession.commit()                

    def checkup(self,filename,maptype_code=None):
        '''
            检查excel中的区域数据，输入的省、市、区县、乡镇是否正确，生产数据是否存在，机构信息是否存在
        '''
        #获取excel对象
        filepath= os.path.join(self.__folder__, filename)
        wb = load_workbook(filepath)

        mydict = {'全部版图':'2913','派送':'2910','自提':'2911'}
        
        for sheetname in mydict:
            print('开始检查%s的%s数据' %(filename,sheetname))
            sheet = wb[sheetname]
            n = 0
            for row in sheet.rows:
                n+=1
                if n > 1 and (row[0].value !=None):
                    tree_root = row[0].value  #树结构的省份根名称
                    org_type = row[1].value   #平台、门店、偏线                    
                    short_name = row[2].value    #机构简称
                    province_name = row[3].value   #省
                    city_name = row[4].value   #市
                    country_name = row[5].value  #县
                    town_name = row[6].value   #镇
                    address = row[7].value   #详细地址
                    #remark = row[8].value  #备注
                    remark=' '
                    
                    #检查地址
                    standard_addr_code_list = self.query_addr_detail(province_name,city_name,country_name,town_name)
                    if len(standard_addr_code_list)==0:
                        log('机构%s没有获取到对应的行政区域编码' %short_name)
                        remark = '%s地址不正确' % remark
                    elif len(standard_addr_code_list)==1:
                        pass
                    else:
                        log('机构%s,%s,%s,%s,%s,匹配多个行政区域编码%s' %(short_name,province_name,city_name,country_name,town_name,standard_addr_code_list)) 
                        remark = '匹配多个行政区域编码%s' % standard_addr_code_list

                    #检查机构简称
                    org_code_list = self.query_org(short_name)
                    if len(org_code_list)==0:
                        log('%s没有机构数据'%short_name)
                        remark = '没有机构数据,%s' % remark
                    elif  len(org_code_list)==1:
                        pass
                    else:
                        log('%s匹配多条机构数据'%short_name)
                        remark = '匹配多条机构数据,%s' % (remark, org_code_list)    

                    #检查生产数据
                    station_list=[]
                    if maptype_code==None:                        
                        station_list = self.query_station(short_name,mydict[sheetname])
                    else:
                        station_list = self.query_station(short_name,maptype_code)
                    
                    if len(station_list)==0:
                        log('%s没有生产数据'%short_name)
                        remark = '没有生产数据,%s' % remark
                    elif  len(station_list)==1:
                        pass
                    else:
                        log('%s匹配多条生产数据'%short_name)
                        remark = '%s匹配%s条生产数据' % (remark,len(station_list))                       

                    #写入错误信息
                    row[8].value = remark
                    
        #保存
        wb.save(filepath)    

    def query_adcode(self, province_name):
        '''
            根据省份名称获取对应的标准地址码
        '''
        adcode=None
        results = addressSession.query(StandardAddr).filter(StandardAddr.level=='province').filter(StandardAddr.name.like(province_name+'%'))
        if results.count()==1:
             adcode=results[0].adcode
        addressSession.close()
        return adcode

    def update_t_area(self):
        '''
            更新t_area表
        '''
        results = gisSession.query(Area).filter(Area.area_type=='STANDARD_ADDRESS')
        for r in results:
            adcode = self.query_adcode(r.area_name)
            r.conn_entity =adcode
            print(r.area_name,adcode)
        gisSession.commit()
                    
        
    def run(self,filename,search_maptype=None):
        '''
            根据指定excel文件中的机构数据和
        '''
        self.todata('收件', filename,search_maptype)  #全部版图的数据
        self.todata('派送',filename,search_maptype) #派送版图的数据
        self.todata('自提',filename,search_maptype) #自提版图的数据

# 程序主入口
if __name__ == '__main__':
    # 实例化执行
    print('开始执行……')
    t= TransferGisData()
    #初始化t_area表的树形结构
    #t.initData('自提')
    #t.initData('派送')
    #t.initData('全部')

    #将旧的制图数据转移到GIS系统中
    #t.run('平台机构数据.xlsx')
    #t.run('加盟门店数据.xlsx')
    #t.run('直营门店数据.xlsx')
    #t.run('偏线机构数据.xlsx','2912')
    #t.run('门店补充数据.xlsx')
    #t.run('平台补充数据.xlsx')
    
    #健康度检查
    #t.checkup('平台机构数据.xlsx')
    #t.checkup('偏线机构数据.xlsx','2912')
    #t.checkup('加盟门店数据.xlsx')
    #t.checkup('直营门店数据.xlsx')
    #t.checkup('门店补充数据.xlsx')
    #t.checkup('平台补充数据.xlsx')
    print('game over')
