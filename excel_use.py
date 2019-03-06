# coding:utf-8
from openpyxl import load_workbook
import config

#获取列数据，需要传入excel名，sheet名，列名
def read_data(filePath,sheetName,list):
    wb = load_workbook(filePath)
    sheet = wb[sheetName]
    a = []
    for i in sheet[list]:
        if i.value != sheet[list+"1" ].value:
            if i.value != None:
                a.append(i.value)
    return a

def write_data(filePath,sheetName,data,response_params):
    wb = load_workbook(filePath)
    sheet = wb[sheetName]
    for i in range(1, len(data) + 1):
        sheet[response_params+"%d" % (i + 1)].value = data[i - 1]
    wb.save(filePath )

if __name__ == "__main__":
    filePath = config.filePath
    sheetName = config.sheetName
    url = config.url
    requests_json = config.requests_json

    data_url = read_data(filePath,sheetName,url)
    print(data_url)

    data_requests_json = read_data(filePath, sheetName, requests_json)
    print(data_requests_json)
    for x, y in zip(data_url, data_requests_json):
        print(x,y)
